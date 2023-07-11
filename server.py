import json
import os
from google.cloud import dialogflow
from flask import Flask, request, jsonify

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'firstagent-huvu-5a35a8485999.json'
def detect_intent_texts(text):
    project_id = 'firstagent-huvu'
    session_id = 'term'
    language_code = 'ko'
    
    # SessionsClient 객체 생성
    session_client = dialogflow.SessionsClient()

    # session path 설정
    session = session_client.session_path(project_id, session_id)

    # text_input 설정
    text_input = dialogflow.TextInput(text=text, language_code=language_code)

    # query_input 설정
    query_input = dialogflow.QueryInput(text=text_input)
    
    # detect_intent 호출
    response = session_client.detect_intent(session=session, query_input=query_input)

    print(response)

detect_intent_texts('COA에 대해 알려줘')


'''import json
from openpyxl import load_workbook

# Load the workbook
wb = load_workbook(r"C:\Users\rnwkv\OneDrive\바탕 화면\이것저것\ex1.xlsx", data_only=True)
ws = wb.active

data = []
for row in ws.iter_rows(min_row=2, values_only=True):
    # Include only the necessary columns in each row
    necessary_row = [str(row[i]).strip() for i in range(len(row)) if i not in [2, 5, 6]]  # Ignoring the 3rd, 6th, and 7th column
    data.append(necessary_row)

# 결과를 저장할 리스트 초기화
entities = []

for row in data:
    value = row[1]
    synonyms = [row[1], row[2], row[2].upper()]

    # 엔티티에 값을 추가
    if 'NONE' not in synonyms:  # Ignore if there is a None in synonyms
        entities.append({"value": value, "synonyms": synonyms})

# 결과를 텍스트 파일에 출력
with open('output.txt', 'w', encoding='utf-8') as f:
    f.write('[\n')  # Open the list
    for i, entity in enumerate(entities):
        line_str = json.dumps(entity, ensure_ascii=False)  # Convert the dictionary to a JSON string without escaping unicode
        line_str = line_str.replace('_x000D_', '')  # Replace '_x000D_' with nothing
        if i != len(entities) - 1:  # If it's not the last entity, add a comma at the end
            f.write(line_str + ',\n')  # Write to the file
        else:  # If it's the last entity, do not add a comma
            f.write(line_str + '\n')  # Write to the file
    f.write(']')  # Close the list
'''